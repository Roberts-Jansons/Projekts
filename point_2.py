import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
import time
import sys


    
def get_word_definition(word, dictionary_number):
    driver = webdriver.Chrome()  # Make sure to use the appropriate driver for your browser

    try:
        word=word.replace(" ","-")
        links=["https://www.dictionary.com/browse/", "https://www.collinsdictionary.com/dictionary/english/", "https://www.oxfordlearnersdictionaries.com/definition/english/"]
        i=dictionary_number
        not_found_defintion=True
        
        while not_found_defintion==True:
            link=links[i]+word
            driver.get(link)
            
            # Find the search input field and enter the word
            #search_box = driver.find_element("id", "q")
            #search_box.clear()
            #search_box.send_keys(word)
            #search_box.send_keys(Keys.RETURN)

            # Find the definition element and retrieve its text

            if i==0:
                try:
                    content_other=driver.find_element("class name", "E17D6zMGNMyhZ9DoRo9R")
                    i+=1
                    not_found_defintion=False
                except:
                    definition_element=driver.find_element("class name", "ESah86zaufmd2_YPdZtq")
                    definition=definition_element.text # Dictionary
                    definition=definition.replace(".","")
                    not_found_defintion=False   
                    return definition
                
            elif i==1:
                try:
                    content=driver.find_element("class name", "cB")
                    i+=1
                    not_found_defintion=False
                except:
                    definition_element=driver.find_element("class name", "def")
                    definition=definition_element.text # no Collins Online Dictionary
                    not_found_defintion=False   
                    return definition
                
            elif i==2:
                try:
                    content=driver.find_element("class name", "results")
                    not_found_defintion=False
                except:
                    definition_element=driver.find_element("class name", "def")
                    definition=definition_element.text # no Oxford Learner's Dictionary
                    not_found_defintion=False   
                    return definition
        

    finally:
        #Close the browser window
        #driver.quit()
        idk=1

words_to_search=["unedited1", "unedited2", "unedited3"]
def manually():
    words_to_search.clear()
    print("Ieraksti stop, lai beigtu.")
    while True:
        adding_word=input()
        if adding_word=="stop":
            break
        adding_word=adding_word.lower()
        words_to_search.append(adding_word)
        
def upload():
    words_to_search.clear()

    print("Nosacījumi, lai vaŗētu izmantot šo veidu:")
    print("  1. Ir iespējams atvērt failu kurš atrodas šajā paša mapē, ko programmas kods.")
    print("  2. Vārdiem / terminime jābūt angliski.")
    print("  3. Šiem terminiem jābūt tikkai kollonā B, sākot no pozīcijas B3.")
    print("Piezīme, ja neatbilst kāds no kritērijiem, ieraksti Stop")
    print("")
    file=input("Kāds excel fails jāatver? (Piezīme: tam jāatrodas tajā pašā mapē kā šis fails): ")+".xlsx"
    if file=="Stop.xlsx":
        print("Neatblist kritērijiem, lietotājs apturēja kodu pirms tā uzsākšanās.")
        sys.exit()


    print("Atver excel failu ", file)
    #file="manually_testam.xlsx"

    wb=load_workbook(file)
    ws=wb.active
    max_row=ws.max_row
    #start=54 # (preformance bonus)
    #start=55 # working
    #start=100
    
    for row in range(3,max_row+1):
        word=(ws['B' + str(row)].value)
        if type(word) is None:
            break
        word=word.lower()
        words_to_search.append(word)
        
    
   
def main():
    choice=input("Kāda veidā gribi atrast vārdu definīcjas? Ja manuāli, ievadi input, bet, ja caur excel dokumentu - upload. ")
    #choice=input("if want to write manually ievadi input, but, if want to use file - upload. ")
    
    if choice=="input" or choice=="upload":
        if choice=="input":
            print("manuāls teksts")
            manually()
        elif choice=="upload":
            print("no dokumenta")
            upload()

        print("")
        print("Jānolasa un jaatrod šādu vārdu definīcijas:")
        print(words_to_search)
        print("")
           
        # Create a new Excel workbook and add a worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Word Definitions"

        # Add headers to the worksheet
        worksheet.append(["Word", "Definition"])

        words_and_definitions=[["to be started", "to be started"]]
        for number_of_dictionary in range(0, 3):
            
            if number_of_dictionary>0:
                #print("Tagad nolasa no ",number_of_dictionary, ". vārdnīcas")                
                for number in range(0, len(words_to_search)):
                    if type(words_and_definitions[number][1]) is not str:
                        #print("Vārdam ", words_and_definitions[number][0]," nav definīcijas")
                        definition=get_word_definition(words_and_definitions[number][0], number_of_dictionary)
                        if type(definition) is str:
                            definition=definition.replace(".","")
                        words_and_definitions[number][1]=definition

            else:
                words_and_definitions.clear()                
                for word in words_to_search:
                    word=word.replace(" ","-")
                    definition = get_word_definition(word, 0)
                    word_and_definition=[word, definition]
                    words_and_definitions.append(word_and_definition)

        print()
        for complete in range(0, len(words_to_search)):
            worksheet.append([words_and_definitions[complete][0], words_and_definitions[complete][1]])
            print(words_and_definitions[complete][0],"---",words_and_definitions[complete][1])
        workbook.save("word_definitions.xlsx")
            
       
    else:
        print("Incorrect input")

    
if __name__ == "__main__":
    main()